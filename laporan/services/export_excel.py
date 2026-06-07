from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from laporan.services.rekap import format_tanggal_short


def generate_excel(queryset) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Harian"

    headers = [
        "No",
        "Nama Siswa",
        "NIS",
        "Jurusan",
        "Uraian Pekerjaan",
        "Hasil / Progress",
        "Status",
        "Tanggal",
        "Jumlah Foto",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for index, laporan in enumerate(queryset, start=1):
        ws.append(
            [
                index,
                laporan.siswa.nama_lengkap,
                laporan.siswa.nis,
                laporan.siswa.jurusan.kode,
                laporan.uraian_pekerjaan,
                laporan.hasil_progress,
                laporan.status,
                format_tanggal_short(laporan.tanggal),
                laporan.jumlah_foto,
            ]
        )

    for row in ws.iter_rows(min_row=2, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
